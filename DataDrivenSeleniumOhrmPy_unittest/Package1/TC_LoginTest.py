import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook


class LoginTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # Set up the WebDriver
        cls.driver = webdriver.Chrome()
        cls.driver.maximize_window()
        cls.base_url = "https://opensource-demo.orangehrmlive.com"
        cls.driver.implicitly_wait(10)

    @classmethod
    def tearDownClass(cls):
        # Quit the WebDriver after all tests
        cls.driver.quit()

    def login(self, username, password):
        """
        Helper method to perform login.
        :param username: Username for login
        :param password: Password for login
        """
        self.driver.get(self.base_url)
        self.driver.find_element(By.NAME, "username").send_keys(username)
        self.driver.find_element(By.NAME, "password").send_keys(password)
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()

    def logout(self):
        """
        Helper method to log out from the application.
        """
        try:
            user_dropdown = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//span[@class='oxd-userdropdown-tab']")))
            user_dropdown.click()

            logout_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//a[text()='Logout']")))
            logout_button.click()

            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, "username")))
        
        except Exception as e:
            self.fail(f"Logout failed due to: {e}")


    def is_dashboard_displayed(self):
        """
        Helper method to check if the Dashboard is displayed.
        :return: True if Dashboard is displayed, False otherwise
        """
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//h6[text()='Dashboard']"))
            )
            return True
        except:
            return False

    def read_login_data(self, file_path, sheet_name):
        """
        Read login data from an Excel file.
        :param file_path: Path to the Excel file
        :param sheet_name: Name of the sheet in the Excel file
        :return: List of tuples containing username and password
        """
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []

        for row in range(2, sheet.max_row + 1):  # Skip header row
            username = sheet.cell(row, 1).value
            password = sheet.cell(row, 2).value
            expected_result = sheet.cell(row, 3).value  # Expected result (e.g., "Pass" or "Fail")
            data.append((username, password, expected_result))

        return data

    def test_login(self):
        """Data-driven test for login."""
        # Path to the Excel file
        file_path = "E:\Rifat\Project\SQA\OrangeHRM\data.xlsx"
        sheet_name = "login"

        # Read login data from Excel
        login_data = self.read_login_data(file_path, sheet_name)

        for username, password, expected_result in login_data:
            with self.subTest(username=username, password=password):
                self.login(username, password)
                if expected_result == "Pass":
                    self.assertTrue(self.is_dashboard_displayed(), f"Login failed for valid credentials: {username}")
                    self.logout()
                else:
                    error_message = self.driver.find_element(By.XPATH, "//p[contains(@class, 'oxd-alert-content-text')]").text
                    self.assertEqual(error_message, "Invalid credentials", f"Unexpected result for: {username}")


if __name__ == "__main__":
    unittest.main()
